"""
create_referentiel_pep.py — ScreenEdge Africa
1. Crée la table referentiel_pep dans compliance_db (si elle n'existe pas)
2. Importe les données depuis PEP_Referentiel_Pays_ScreenEdge_Africa_v11.xlsx
   → Feuille "Référentiel PEP par pays"  (définitions officielles)
   → Feuille "Statut GAFI & Couverture"  (statuts + priorités)

Usage : python create_referentiel_pep.py
"""

import os
import sys
import re

import openpyxl
from db_utils import get_pg_conn

sys.stdout.reconfigure(encoding="utf-8")

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          "PEP_Referentiel_Pays_ScreenEdge_Africa_v11.xlsx")

# ── DDL ────────────────────────────────────────────────────────────────────────

DDL = """
CREATE TABLE IF NOT EXISTS referentiel_pep (
    id          SERIAL PRIMARY KEY,
    region      TEXT NOT NULL CHECK (region IN ('maghreb', 'uemoa', 'autre')),
    pays        TEXT NOT NULL,
    code_iso    CHAR(2) NOT NULL,
    loi_ref     TEXT,
    def_pep     TEXT,
    statut_gafi TEXT NOT NULL DEFAULT 'clean'
                     CHECK (statut_gafi IN ('clean', 'liste_grise', 'liste_noire')),
    vigilance   TEXT NOT NULL DEFAULT 'standard'
                     CHECK (vigilance IN ('standard', 'renforcee', 'maximale')),
    autorite    TEXT,
    source_url  TEXT,
    notes       TEXT,
    updated_at  TIMESTAMP WITH TIME ZONE DEFAULT NOW(),
    UNIQUE (code_iso)
);

CREATE INDEX IF NOT EXISTS idx_ref_pep_code_iso ON referentiel_pep (code_iso);
CREATE INDEX IF NOT EXISTS idx_ref_pep_statut   ON referentiel_pep (statut_gafi);
"""


# ── Normalisation des valeurs ──────────────────────────────────────────────────

REGION_MAP = {
    "maghreb":           "maghreb",
    "afrique de l'ouest": "uemoa",
    "afrique de louest":  "uemoa",
    "uemoa":             "uemoa",
    "autre":             "autre",
    "autres":            "autre",
}


def norm_region(val: str) -> str:
    if not val:
        return "autre"
    return REGION_MAP.get(val.strip().lower(), "autre")


def norm_statut_gafi(val: str) -> str:
    if not val:
        return "clean"
    v = val.lower()
    if "liste noire" in v or "liste_noire" in v:
        return "liste_noire"
    if "liste grise" in v or "liste_grise" in v or "⚠️" in val:
        return "liste_grise"
    return "clean"


def norm_vigilance(statut: str, notes: str) -> str:
    """Déduit la vigilance depuis le statut GAFI et les notes."""
    if statut == "liste_noire":
        return "maximale"
    if statut == "liste_grise":
        return "renforcee"
    # Pays clean mais vigilance renforcée recommandée (ex: Libye, Guinée, Mali, Niger)
    if notes and re.search(r"renforcée? (recommandée?|obligatoire)", notes, re.IGNORECASE):
        return "renforcee"
    return "standard"


def clean(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ── Lecture de l'Excel ─────────────────────────────────────────────────────────

def lire_excel() -> list[dict]:
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # Feuille 1 : Référentiel PEP par pays (colonnes A→J, lignes 4+)
    ws_ref = wb["Référentiel PEP par pays"]
    ref_rows = {}
    for row in ws_ref.iter_rows(min_row=4, values_only=True):
        region, pays, code_iso, def_pep, loi_ref, _, autorite, source_url, statut_raw, notes = row
        if not pays or not code_iso:
            continue
        code = str(code_iso).strip().upper()
        statut = norm_statut_gafi(str(statut_raw or ""))
        notes_str = clean(notes)
        ref_rows[code] = {
            "region":      norm_region(str(region or "")),
            "pays":        str(pays).strip(),
            "code_iso":    code,
            "def_pep":     clean(def_pep),
            "loi_ref":     clean(loi_ref),
            "autorite":    clean(autorite),
            "source_url":  clean(source_url),
            "statut_gafi": statut,
            "vigilance":   norm_vigilance(statut, notes_str),
            "notes":       notes_str,
        }

    # Feuille 2 : Statut GAFI & Couverture — complète/corrige statut_gafi + autorite
    ws_gafi = wb["Statut GAFI & Couverture"]
    for row in ws_gafi.iter_rows(min_row=3, values_only=True):
        pays_g, region_gafi, statut_g, organisme, source_pep, priorite = row
        if not pays_g:
            continue
        # Retrouve le code_iso par correspondance du nom de pays
        code_found = None
        for code, data in ref_rows.items():
            if data["pays"].lower() == str(pays_g).strip().lower():
                code_found = code
                break
        if not code_found:
            continue
        # Met à jour statut + vigilance depuis cette feuille (plus concise)
        statut = norm_statut_gafi(str(statut_g or ""))
        ref_rows[code_found]["statut_gafi"] = statut
        ref_rows[code_found]["vigilance"]   = norm_vigilance(statut, ref_rows[code_found].get("notes", ""))
        if organisme and not ref_rows[code_found]["autorite"]:
            ref_rows[code_found]["autorite"] = str(organisme).strip()

    print(f"Excel lu : {len(ref_rows)} pays trouvés")
    return list(ref_rows.values())


# ── Insertion / Upsert ─────────────────────────────────────────────────────────

UPSERT_SQL = """
INSERT INTO referentiel_pep
    (region, pays, code_iso, loi_ref, def_pep, statut_gafi, vigilance, autorite, source_url, notes)
VALUES
    (%(region)s, %(pays)s, %(code_iso)s, %(loi_ref)s, %(def_pep)s,
     %(statut_gafi)s, %(vigilance)s, %(autorite)s, %(source_url)s, %(notes)s)
ON CONFLICT (code_iso)
DO UPDATE SET
    region      = EXCLUDED.region,
    pays        = EXCLUDED.pays,
    loi_ref     = EXCLUDED.loi_ref,
    def_pep     = EXCLUDED.def_pep,
    statut_gafi = EXCLUDED.statut_gafi,
    vigilance   = EXCLUDED.vigilance,
    autorite    = EXCLUDED.autorite,
    source_url  = EXCLUDED.source_url,
    notes       = EXCLUDED.notes,
    updated_at  = NOW();
"""


def run():
    print("\n" + "=" * 60)
    print("create_referentiel_pep.py — ScreenEdge Africa")
    print("=" * 60)

    rows = lire_excel()

    with get_pg_conn() as conn:
        with conn.cursor() as cur:
            # Création de la table
            print("\n→ Création de la table referentiel_pep...")
            cur.execute(DDL)
            print("  Table et index créés (ou déjà existants)")

            # Upsert ligne par ligne avec retour clair
            print("\n→ Import des données...")
            for row in rows:
                cur.execute(UPSERT_SQL, row)
                print(f"  ✓ {row['pays']} ({row['code_iso']}) — {row['statut_gafi']} / {row['vigilance']}")

        conn.commit()

    print(f"\n✅ Import terminé : {len(rows)} pays insérés/mis à jour dans compliance_db.referentiel_pep")
    print("=" * 60)


if __name__ == "__main__":
    run()
