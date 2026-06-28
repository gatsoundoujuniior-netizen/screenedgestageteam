"""
excel_to_json.py — ScreenEdge Africa
Convertit PEP_Referentiel_Pays_ScreenEdge_Africa_v11.xlsx
en referentiel_pep.json structuré pour le LLM.
"""

import sys, json, re
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from dotenv import load_dotenv
from langchain_groq import ChatGroq

load_dotenv(override=True)

EXCEL_PATH = "../PEP_Referentiel_Pays_ScreenEdge_Africa_v11.xlsx"
OUTPUT_PATH = "referentiel_pep.json"

llm = ChatGroq(model="llama-3.1-8b-instant", temperature=0)

PROMPT_EXTRACTION = """Tu es un expert AML/PPE. Extrait les informations suivantes du texte juridique ci-dessous et réponds UNIQUEMENT en JSON valide.

TEXTE :
{def_pep}

Réponds avec ce JSON :
{{
  "fonctions_pep": ["liste", "des", "fonctions", "PEP", "du", "pays"],
  "famille_incluse": true ou false,
  "proches_associes": true ou false,
  "duree_ex_pep": "permanente ou X mois/ans ou non précisée",
  "reevaluation": "fréquence de réévaluation ou non précisée"
}}

RÈGLES :
- fonctions_pep = liste courte et claire des titres/fonctions qui qualifient comme PEP (max 12 items)
- Chaque item doit être un titre clair ex: "Chef d'État", "Ministre", "Parlementaire"
- Ne pas inclure famille ni proches dans fonctions_pep — ils ont leurs propres champs
- Si le texte dit "exercent ou ont exercé" → duree_ex_pep = "permanente"
- Répondre UNIQUEMENT en JSON, aucun texte avant ou après"""


def extraire_statut_gafi(texte: str) -> str:
    """Extrait le statut GAFI depuis la cellule Notes."""
    if not texte:
        return "clean"
    t = texte.lower()
    if "liste noire" in t:
        return "liste_noire"
    if "liste grise" in t:
        return "liste_grise"
    return "clean"


def extraire_vigilance(statut_gafi: str, notes: str) -> str:
    if statut_gafi == "liste_noire":
        return "maximale"
    if statut_gafi == "liste_grise":
        return "renforcee"
    if notes and "renforcée" in notes.lower():
        return "renforcee"
    return "standard"


def appeler_llm(def_pep: str) -> dict:
    """Utilise Groq pour extraire les fonctions PEP depuis le texte."""
    try:
        resp = llm.invoke(PROMPT_EXTRACTION.format(def_pep=def_pep[:3000]))
        content = resp.content.strip()
        if "```json" in content:
            content = content.split("```json")[1].split("```")[0].strip()
        elif "```" in content:
            content = content.split("```")[1].split("```")[0].strip()
        start = content.find("{")
        end   = content.rfind("}") + 1
        return json.loads(content[start:end])
    except Exception as e:
        print(f"  Erreur LLM : {e}")
        return {
            "fonctions_pep": ["Chef d'État", "Ministre", "Parlementaire",
                              "Membre juridiction supérieure", "Gouverneur Banque Centrale",
                              "Ambassadeur", "Officier supérieur", "Dirigeant entreprise publique",
                              "Haut responsable parti politique"],
            "famille_incluse": True,
            "proches_associes": True,
            "duree_ex_pep": "permanente",
            "reevaluation": "non précisée"
        }


def main():
    print("Lecture Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # Feuille 1 — Référentiel PEP par pays
    ws_ref    = wb["Référentiel PEP par pays"]
    # Feuille 3 — Statut GAFI & Couverture
    ws_gafi   = wb["Statut GAFI & Couverture"]

    # Charger statuts GAFI depuis feuille 3 (plus fiable que les notes)
    statuts_gafi = {}
    for row in ws_gafi.iter_rows(min_row=3, values_only=True):
        if row[0] and row[2]:
            pays_nom = str(row[0]).strip()
            statut   = str(row[2]).strip().lower()
            if "liste noire" in statut:
                statuts_gafi[pays_nom] = "liste_noire"
            elif "liste grise" in statut:
                statuts_gafi[pays_nom] = "liste_grise"
            else:
                statuts_gafi[pays_nom] = "clean"

    referentiel = []

    print("\nTraitement des pays...")
    for row in ws_ref.iter_rows(min_row=4, values_only=True):
        # Colonnes : Région, Pays, Code ISO, Définition PEP, Loi référence,
        #            Liste officielle, Organisme, URL, Statut GAFI, Notes
        region, pays, code_iso, def_pep, loi_ref, liste_off, organisme, url, _, notes = (
            row[i] if i < len(row) else None for i in range(10)
        )

        if not pays or not code_iso:
            continue

        pays      = str(pays).strip()
        code_iso  = str(code_iso).strip().upper()
        def_pep   = str(def_pep).strip() if def_pep else ""
        loi_ref   = str(loi_ref).strip() if loi_ref else ""
        url       = str(url).strip() if url else ""
        notes     = str(notes).strip() if notes else ""

        print(f"  {code_iso} — {pays}...")

        # Statut GAFI depuis feuille 3 (priorité) ou extraction texte
        statut_gafi = statuts_gafi.get(pays, extraire_statut_gafi(notes))
        vigilance   = extraire_vigilance(statut_gafi, notes)

        # Extraction fonctions via LLM
        extraction = appeler_llm(def_pep) if def_pep else {
            "fonctions_pep": [],
            "famille_incluse": False,
            "proches_associes": False,
            "duree_ex_pep": "non précisée",
            "reevaluation": "non précisée"
        }

        referentiel.append({
            "pays":              pays,
            "code_iso":          code_iso,
            "region":            str(region).strip() if region else "",
            "statut_gafi":       statut_gafi,
            "vigilance":         vigilance,
            "loi_reference":     loi_ref,
            "organisme":         str(organisme).strip() if organisme else "",
            "source_url":        url,
            "fonctions_pep":     extraction.get("fonctions_pep", []),
            "famille_incluse":   extraction.get("famille_incluse", True),
            "proches_associes":  extraction.get("proches_associes", True),
            "duree_ex_pep":      extraction.get("duree_ex_pep", "permanente"),
            "reevaluation":      extraction.get("reevaluation", "non précisée"),
            "def_pep_complet":   def_pep,  # texte original conservé en référence
        })

    # Sauvegarde JSON
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(referentiel, f, ensure_ascii=False, indent=2)

    print(f"\n✅ {len(referentiel)} pays traités → {OUTPUT_PATH}")
    print("\nAperçu (Sénégal) :")
    for p in referentiel:
        if p["code_iso"] == "SN":
            print(json.dumps(p, ensure_ascii=False, indent=2))
            break


if __name__ == "__main__":
    main()
